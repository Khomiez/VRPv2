"""
Convert the VRP input Excel template to JSON format for the solver.
"""
import json
import sys
import openpyxl


def convert_template_to_json(excel_path: str, output_path: str = None):
    """
    Convert VRP Excel template to JSON format.

    Args:
        excel_path: Path to the Excel template file
        output_path: Path for output JSON file (default: same name with .json extension)
    """
    if output_path is None:
        output_path = excel_path.rsplit('.', 1)[0] + '.json'

    print(f"Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    # ==================== Read Settings ====================
    ws_settings = wb["Settings"]
    settings = {}
    for row in ws_settings.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1] is not None:
            settings[row[0]] = row[1]

    depot_id = int(settings.get("Depot Node ID", 1))
    depot_name = str(settings.get("Depot Name", "Depot"))

    # ==================== Read Nodes ====================
    ws_nodes = wb["Nodes"]
    nodes = []

    for row in ws_nodes.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Stop at empty row
            break

        node = {
            "id": int(row[0]),
            "name": str(row[1]) if row[1] else f"Node {row[0]}",
            "general_demand": float(row[2]) if row[2] is not None else 0.0,
            "recycle_demand": float(row[3]) if row[3] is not None else 0.0,
        }
        # Store node type for later use (not in final JSON but useful for validation)
        node_type = str(row[4]).lower() if row[4] else "customer"
        node["_type"] = node_type
        nodes.append(node)

    # Sort nodes by ID
    nodes.sort(key=lambda x: x["id"])
    num_nodes = len(nodes)

    print(f"Found {num_nodes} nodes")

    # Validate depot
    depot_found = False
    checkpoint_found = False
    for node in nodes:
        if node["id"] == depot_id:
            depot_found = True
        if node.get("_type") == "checkpoint":
            checkpoint_found = True

    if not depot_found:
        print(f"WARNING: Depot node with ID {depot_id} not found in nodes!")
    if not checkpoint_found:
        print("WARNING: No checkpoint node found! At least one node should have type 'checkpoint'")

    # Remove temporary _type field
    for node in nodes:
        del node["_type"]

    # ==================== Read Vehicles ====================
    ws_vehicles = wb["Vehicles"]
    vehicles = []

    for row in ws_vehicles.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Stop at empty row
            break

        vehicle = {
            "type": str(row[0]),
            "general_capacity": float(row[1]) if row[1] is not None else 0.0,
            "recycle_capacity": float(row[2]) if row[2] is not None else 0.0,
            "fixed_cost": float(row[3]) if row[3] is not None else 0.0,
            "fuel_cost_per_distance": float(row[4]) if row[4] is not None else 0.0,
        }
        vehicles.append(vehicle)

    print(f"Found {len(vehicles)} vehicle types")

    # ==================== Read Distance Matrix ====================
    ws_distance = wb["Distance_Matrix"]

    # Read header row to get node ID order
    header_row = list(ws_distance.iter_rows(min_row=1, max_row=1, min_col=2, values_only=True))[0]
    matrix_node_ids = [int(x) for x in header_row if x is not None]
    matrix_size = len(matrix_node_ids)

    print(f"Distance matrix size: {matrix_size}x{matrix_size}")

    if matrix_size != num_nodes:
        print(f"WARNING: Distance matrix size ({matrix_size}) doesn't match number of nodes ({num_nodes})")

    # Read distance values
    distance_matrix = []
    for row_idx, row in enumerate(ws_distance.iter_rows(min_row=2, max_row=matrix_size+1, min_col=2, max_col=matrix_size+1, values_only=True)):
        row_values = []
        for val in row:
            if val is None:
                row_values.append(0.0)
            else:
                row_values.append(float(val))
        distance_matrix.append(row_values)

    # Validate matrix is symmetric
    for i in range(len(distance_matrix)):
        for j in range(len(distance_matrix[i])):
            if abs(distance_matrix[i][j] - distance_matrix[j][i]) > 0.01:
                print(f"WARNING: Distance matrix not symmetric at [{i+1}][{j+1}]: {distance_matrix[i][j]} != {distance_matrix[j][i]}")

    # ==================== Build Output JSON ====================
    total_general = sum(n["general_demand"] for n in nodes)
    total_recycle = sum(n["recycle_demand"] for n in nodes)

    output = {
        "depot": {
            "id": depot_id,
            "name": depot_name
        },
        "num_nodes": num_nodes,
        "nodes": nodes,
        "vehicles": vehicles,
        "distance_matrix": distance_matrix,
        "summary": {
            "total_general_demand": total_general,
            "total_recycle_demand": total_recycle,
            "num_vehicles": len(vehicles)
        }
    }

    # Write JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nJSON output saved to: {output_path}")
    print(f"\nSummary:")
    print(f"  - Total nodes: {num_nodes}")
    print(f"  - Total general demand: {total_general}")
    print(f"  - Total recycle demand: {total_recycle}")
    print(f"  - Vehicle types: {len(vehicles)}")

    return output


def main():
    if len(sys.argv) < 2:
        print("Usage: python convert_template_to_json.py <excel_file> [output_json]")
        print("\nExample:")
        print("  python convert_template_to_json.py vrp_input_template.xlsx")
        print("  python convert_template_to_json.py my_data.xlsx my_output.json")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    convert_template_to_json(excel_path, output_path)


if __name__ == "__main__":
    main()
